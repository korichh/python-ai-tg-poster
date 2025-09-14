import asyncio
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from local_types import ErrorType
from utils import logger


class PostRepository:
    posts_file = Path("public") / "posts.xlsx"
    _lock: asyncio.Lock = asyncio.Lock()

    @staticmethod
    async def _create_posts_file() -> None:
        async with PostRepository._lock:
            wb = Workbook()
            ws = wb.active
            ws.title = "Posts"
            ws.append(["Date", "Text", "Image Path"])
            wb.save(PostRepository.posts_file)

    @staticmethod
    async def _ensure_posts_file() -> None:
        PostRepository.posts_file.parent.mkdir(parents=True, exist_ok=True)

        if not PostRepository.posts_file.exists():
            await PostRepository._create_posts_file()

        else:
            try:
                async with PostRepository._lock:
                    load_workbook(PostRepository.posts_file)

            except (InvalidFileException, KeyError, OSError):
                await PostRepository._create_posts_file()

    @staticmethod
    async def add_post(date: str, text: str, image_path: str | None) -> None:
        try:
            await PostRepository._ensure_posts_file()

            async with PostRepository._lock:
                wb = load_workbook(PostRepository.posts_file)
                ws = wb.active

                ws.append([date, text, image_path if image_path else ""])
                wb.save(PostRepository.posts_file)

        except Exception as exc:
            logger.error(f"{ErrorType.POST_REPOSITORY_ERROR.value}: {str(exc)}")

            return None
